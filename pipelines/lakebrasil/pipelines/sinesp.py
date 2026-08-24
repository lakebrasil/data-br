"""SINESP — eventos de segurança pública por município → indicadores_serie.

Carrega `s3://data-br-raw/sinesp/raw/basededadosvde.zip` (10 anos
2015-2024 em CSVs anuais, ~600 MB descompactado). Cada CSV tem 1 row
por (uf, município, evento, data_referencia, agente, arma, ...) =
ocorrência individual. Agregamos para nível município-mês:
  GROUP BY (uf, municipio, evento, ano-mes)
  SUM(total_vitima) → valor

Eventos padrão (~25 categorias): homicídio doloso, feminicídio,
roubo, latrocínio, suicídio, mortes no trânsito, etc.

Mapeamento → `data_br.indicadores_serie`:
  ibge_code     ← resolve_ibge(uf, municipio) (slug match)
  fonte         ← 'sinesp'
  indicador_id  ← `sinesp.{evento_slug}` (e.g. `sinesp.homicidio_doloso`)
  periodo       ← YYYY-MM
  valor         ← SUM(total_vitima) ou COUNT(*) se total_vitima ausente
  unidade       ← 'ocorrencias'

Source: gov.br/mj/dados-abertos/sinesp.

Uso:
    AWS_PROFILE=<seu-perfil> python -m lakebrasil.pipelines.sinesp --no-fetch
    AWS_PROFILE=<seu-perfil> python -m lakebrasil.pipelines.sinesp --year 2024 --dry-run
"""
from __future__ import annotations

import argparse
import datetime as _dt
import io
import re
import sys
import time
import unicodedata
from collections import defaultdict
from collections.abc import Iterator

import dlt

from lakebrasil.common.args import add_common_args
from lakebrasil.common.enrich import municipios_count, resolve_ibge
from lakebrasil.common.fetch import ensure_fetched
from lakebrasil.common.incremental import loaded_triples
from lakebrasil.common.s3 import get_object_bytes, list_keys
from lakebrasil.pipelines.destinations.s3tables import s3tables_iceberg

import openpyxl

S3_PREFIX = "sinesp/raw/"
# `basededadosvde.zip` (nunca existiu — suposição antiga do docstring)
# não é o que o MJ realmente publica. O real é 1 xlsx por ano em
# gov.br/mj/.../download/dnsp-base-de-dados/bancovde-{ano}.xlsx —
# mesmas colunas (uf, municipio, evento, data_referencia, total_vitima,
# ...) só que já num sheet único ao invés de CSV dentro de um zip.
FILE_RE = re.compile(r"^bancovde-(\d{4})\.xlsx$", re.IGNORECASE)


def _slug(text: str) -> str:
    """Normaliza nome de evento → indicador slug ASCII safe.
    Ex: 'Homicídio doloso' → 'homicidio_doloso'."""
    s = unicodedata.normalize("NFKD", text)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().strip()
    return re.sub(r"[^a-z0-9]+", "_", s).strip("_")[:40]


def _iter_year_xlsx(xlsx_bytes: bytes, ano: int) -> Iterator[dict]:
    """Stream bancovde-{ano}.xlsx → 1 row por (uf, mun, evento, periodo)
    pré-agregado em memória."""
    accum: dict[tuple[str, str, str, str], int] = defaultdict(int)
    rows_seen = 0
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header: list[str] | None = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c else "" for c in row]
            continue
        rows_seen += 1
        r = dict(zip(header, row))
        uf = (r.get("uf") or "").strip().upper()
        mun = (r.get("municipio") or "").strip()
        evento = (r.get("evento") or "").strip()
        data_ref = r.get("data_referencia")
        periodo = (f"{data_ref.year:04d}-{data_ref.month:02d}"
                   if isinstance(data_ref, _dt.datetime) else None)
        if not (uf and mun and evento and periodo):
            continue
        # Use total_vitima quando disponível, senão soma 1 (cada
        # row é 1 ocorrência registrada).
        vitima = r.get("total_vitima")
        try:
            qtd = int(vitima) if vitima not in (None, "") else 1
        except (ValueError, TypeError):
            qtd = 1
        if qtd == 0:
            qtd = 1
        accum[(uf, mun, evento, periodo)] += qtd
    wb.close()
    print(f"  SINESP {ano}: {rows_seen:,} rows lidas → {len(accum):,} (uf,mun,evento,mes) "
          f"agregados", file=sys.stderr)
    skipped_ibge = 0
    for (uf, mun, evento, periodo), qtd in accum.items():
        ibge = resolve_ibge(uf, mun)
        if ibge is None:
            # Schema indicadores_serie requer ibge_code NOT NULL; skipa
            # ocorrências onde slug match falhou (geralmente municípios
            # com grafia não-canônica ou agregados estaduais).
            skipped_ibge += 1
            continue
        yield {
            "ibge_code":     ibge,
            "uf":            uf,
            "fonte":         "sinesp",
            "indicador_id":  f"sinesp.{_slug(evento)}",
            "periodo":       periodo,
            "valor":         float(qtd),
            "valor_texto":   None,
            "unidade":       "ocorrencias",
            "fonte_arquivo": f"bancovde-{ano}.xlsx",
        }
    if skipped_ibge:
        print(f"  SINESP {ano}: skipped {skipped_ibge:,} (uf,mun) sem ibge_match",
              file=sys.stderr)


def _build_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    p.add_argument("--year", type=int, help="YYYY single year (debug).")
    add_common_args(p, table_default="indicadores_serie")
    return p.parse_args()


def main() -> int:
    args = _build_args()

    if not args.no_fetch:
        try:
            ensure_fetched("sinesp_*", refresh=args.refresh)
        except ValueError as e:
            print(f"  [warn] {e}", file=sys.stderr)
            print("  [warn] assumindo s3://data-br-raw/sinesp/raw/ já populado.",
                  file=sys.stderr)

    municipios_count()  # warm dim cache pra resolve_ibge
    already = set() if args.full_refresh else loaded_triples(
        "indicadores_serie", "fonte", "indicador_id", "periodo")

    @dlt.resource(
        name="indicadores_serie",
        primary_key=["ibge_code", "indicador_id", "periodo"],
        write_disposition="append",
    )
    def sinesp_indicadores() -> Iterator[dict]:
        keys = sorted(list_keys(S3_PREFIX))
        ano_to_key: dict[int, str] = {}
        for key in keys:
            name = key.rsplit("/", 1)[-1]
            m = FILE_RE.match(name)
            if m:
                ano_to_key[int(m.group(1))] = key
        if not ano_to_key:
            print(f"  [warn] nenhum bancovde-*.xlsx encontrado em {S3_PREFIX}",
                  file=sys.stderr)
            return
        print(f"  SINESP anos disponíveis: {sorted(ano_to_key)}", file=sys.stderr)
        for ano in sorted(ano_to_key):
            if args.year and ano != args.year:
                continue
            t0 = time.monotonic()
            xlsx_bytes = get_object_bytes(ano_to_key[ano])
            for rec in _iter_year_xlsx(xlsx_bytes, ano):
                if ("sinesp", rec["indicador_id"], rec["periodo"]) in already:
                    continue
                yield rec
            print(f"  SINESP {ano}: pronto em {time.monotonic()-t0:.1f}s",
                  file=sys.stderr)

    if args.dry_run:
        pipe = dlt.pipeline(pipeline_name="sinesp_dryrun", destination="duckdb",
                            dataset_name="staging", dev_mode=True)
        pipe.run(sinesp_indicadores())
        with pipe.sql_client() as c:
            print(c.execute_sql(
                "SELECT indicador_id, count(*), sum(valor) "
                "FROM indicadores_serie GROUP BY indicador_id "
                "ORDER BY count(*) DESC LIMIT 10"))
        return 0

    pipe = dlt.pipeline(pipeline_name="sinesp", destination=s3tables_iceberg)
    info = pipe.run(sinesp_indicadores())
    print(info)
    return 0


if __name__ == "__main__":
    sys.exit(main())

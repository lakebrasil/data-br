"""ANP preços combustível por município → data_br.indicadores_serie.

Carrega os relatórios semanais "TIPO RELATÓRIO: MUNICÍPIO" do Sistema de
Levantamento de Preços (SLP) da ANP:

    s3://data-br-raw/anp/raw/semanal-municipi{o,os}-*.xlsx

Não existe (nunca existiu) um `precos-{produto}-{ano-mes}.csv` publicado
pela ANP — isso era uma suposição do docstring original que nunca bateu
com nenhum arquivo real (o pipeline sempre lia 0 arquivos). O que a ANP
de fato publica em gov.br/anp/.../shlp/semanal/ são XLSX multi-ano já
agregados por (semana, região, estado, município, produto), com colunas:

    DATA INICIAL | DATA FINAL | REGIÃO | ESTADO | MUNICÍPIO | PRODUTO |
    NÚMERO DE POSTOS PESQUISADOS | UNIDADE DE MEDIDA |
    PREÇO MÉDIO REVENDA | DESVIO PADRÃO REVENDA | PREÇO MÍNIMO REVENDA |
    PREÇO MÁXIMO REVENDA | MARGEM MÉDIA REVENDA | COEF DE VARIAÇÃO REVENDA

`ESTADO` vem por extenso sem acento ("SAO PAULO", "MARANHAO") — não UF
— precisa de `ESTADO_TO_UF` antes de `resolve_ibge`.

Reagrupamos as semanas (que já vêm com preço médio pré-calculado por
posto pesquisado) em município-mês-produto via média ponderada pelo
número de postos pesquisados em cada semana:

    GROUP BY (uf, municipio, produto, ano-mes)
    weighted_avg(PREÇO MÉDIO REVENDA, peso=NÚMERO DE POSTOS PESQUISADOS)

Mapeamento → indicadores_serie:
  ibge_code     ← resolve_ibge(uf, municipio)
  fonte         ← 'anp'
  indicador_id  ← `anp.preco_{produto_slug}` (e.g. `anp.preco_diesel`, `anp.preco_gnv`)
  periodo       ← YYYY-MM (mês da DATA INICIAL da semana)
  valor         ← média ponderada de PREÇO MÉDIO REVENDA
  unidade       ← 'R$/litro' | 'R$/m³' (de UNIDADE DE MEDIDA)

Source: gov.br/anp Levantamento de Preços de Combustíveis — Séries
Históricas de Preços — relatório por Município.

Uso:
    AWS_PROFILE=<seu-perfil> python -m lakebrasil.pipelines.anp_precos --no-fetch
"""
from __future__ import annotations

import argparse
import datetime as _dt
import io
import re
import sys
import time
import unicodedata
from collections import defaultdict
from collections.abc import Iterator

import dlt
import openpyxl

from lakebrasil.common.args import add_common_args
from lakebrasil.common.enrich import municipios_count, resolve_ibge
from lakebrasil.common.fetch import ensure_fetched
from lakebrasil.common.incremental import loaded_triples
from lakebrasil.common.s3 import get_object_bytes, list_keys
from lakebrasil.pipelines.destinations.s3tables import s3tables_iceberg

S3_PREFIX = "anp/raw/"
# Nomes reais publicados pela ANP variam entre singular/plural:
# "semanal-municipio-2024-2025.xlsx", "semanal-municipios-2026.xlsx".
FILE_RE = re.compile(r"^semanal-municipi(?:o|os)-.*\.xlsx$")

ESTADO_TO_UF = {
    "ACRE": "AC", "ALAGOAS": "AL", "AMAPA": "AP", "AMAZONAS": "AM",
    "BAHIA": "BA", "CEARA": "CE", "DISTRITO FEDERAL": "DF",
    "ESPIRITO SANTO": "ES", "GOIAS": "GO", "MARANHAO": "MA",
    "MATO GROSSO": "MT", "MATO GROSSO DO SUL": "MS", "MINAS GERAIS": "MG",
    "PARA": "PA", "PARAIBA": "PB", "PARANA": "PR", "PERNAMBUCO": "PE",
    "PIAUI": "PI", "RIO DE JANEIRO": "RJ", "RIO GRANDE DO NORTE": "RN",
    "RIO GRANDE DO SUL": "RS", "RONDONIA": "RO", "RORAIMA": "RR",
    "SANTA CATARINA": "SC", "SAO PAULO": "SP", "SERGIPE": "SE",
    "TOCANTINS": "TO",
}

HEADER_ROW = (
    "DATA INICIAL", "DATA FINAL", "REGIÃO", "ESTADO", "MUNICÍPIO",
    "PRODUTO", "NÚMERO DE POSTOS PESQUISADOS", "UNIDADE DE MEDIDA",
    "PREÇO MÉDIO REVENDA",
)


def _slug(text: str) -> str:
    s = unicodedata.normalize("NFKD", text)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().strip()
    return re.sub(r"[^a-z0-9]+", "_", s).strip("_")[:40]


def _iter_xlsx(s3_key: str) -> Iterator[dict]:
    """Stream o xlsx semanal → agrega pra (uf, mun, produto, ano-mes)
    via média ponderada por nº de postos pesquisados na semana."""
    name = s3_key.rsplit("/", 1)[-1]
    print(f"  ANP {name}: load", file=sys.stderr)
    raw = get_object_bytes(s3_key)
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Acumulador (uf, mun, produto, periodo, unidade) → (soma_ponderada, peso_total)
    accum: dict[tuple[str, str, str, str, str], tuple[float, float]] = defaultdict(
        lambda: (0.0, 0.0)
    )
    header: list[str] | None = None
    rows_seen = 0
    rows_skip = 0
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if row and row[0] == "DATA INICIAL":
                header = [str(c).strip() if c else "" for c in row]
            continue
        rows_seen += 1
        rec = dict(zip(header, row))
        data_inicial = rec.get("DATA INICIAL")
        estado = (rec.get("ESTADO") or "").strip().upper()
        mun = (rec.get("MUNICÍPIO") or "").strip()
        produto = (rec.get("PRODUTO") or "").strip()
        if not (isinstance(data_inicial, _dt.datetime) and estado and mun and produto):
            rows_skip += 1
            continue
        uf = ESTADO_TO_UF.get(estado)
        if uf is None:
            rows_skip += 1
            continue
        preco = rec.get("PREÇO MÉDIO REVENDA")
        n_postos = rec.get("NÚMERO DE POSTOS PESQUISADOS")
        if not isinstance(preco, (int, float)) or not isinstance(n_postos, (int, float)) or n_postos <= 0:
            rows_skip += 1
            continue
        unidade = (rec.get("UNIDADE DE MEDIDA") or "").strip() or "R$/litro"
        periodo = f"{data_inicial.year:04d}-{data_inicial.month:02d}"
        key = (uf, mun, produto, periodo, unidade)
        soma, peso = accum[key]
        accum[key] = (soma + preco * n_postos, peso + n_postos)
    print(f"  ANP {name}: rows={rows_seen:,} skip={rows_skip:,} → "
          f"{len(accum):,} chaves (uf,mun,produto,periodo)", file=sys.stderr)

    skipped_ibge = 0
    for (uf, mun, produto, periodo, unidade), (soma, peso) in accum.items():
        ibge = resolve_ibge(uf, mun)
        if ibge is None:
            skipped_ibge += 1
            continue
        yield {
            "ibge_code":     ibge,
            "uf":            uf,
            "fonte":         "anp",
            "indicador_id":  f"anp.preco_{_slug(produto)}",
            "periodo":       periodo,
            "valor":         soma / peso,
            "valor_texto":   None,
            "unidade":       unidade,
            "fonte_arquivo": name,
        }
    if skipped_ibge:
        print(f"  ANP {name}: skipped {skipped_ibge:,} (uf,mun) sem ibge_match",
              file=sys.stderr)


def _build_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    add_common_args(p, table_default="indicadores_serie")
    return p.parse_args()


def main() -> int:
    args = _build_args()

    if not args.no_fetch:
        try:
            ensure_fetched("anp_precos_*", refresh=args.refresh)
        except ValueError:
            print("  [warn] assumindo s3://data-br-raw/anp/raw/semanal-municipi*.xlsx já populado.",
                  file=sys.stderr)

    municipios_count()  # warm dim
    already = set() if args.full_refresh else loaded_triples(
        "indicadores_serie", "fonte", "indicador_id", "periodo")

    @dlt.resource(
        name="indicadores_serie",
        primary_key=["ibge_code", "indicador_id", "periodo"],
        write_disposition="append",
    )
    def anp_precos() -> Iterator[dict]:
        for key in sorted(list_keys(S3_PREFIX)):
            name = key.rsplit("/", 1)[-1]
            if not FILE_RE.match(name):
                continue
            t0 = time.monotonic()
            n = 0
            for rec in _iter_xlsx(key):
                if ("anp", rec["indicador_id"], rec["periodo"]) in already:
                    continue
                n += 1
                yield rec
            print(f"  ANP {name}: {n:,} registros em {time.monotonic()-t0:.1f}s",
                  file=sys.stderr)

    if args.dry_run:
        pipe = dlt.pipeline(pipeline_name="anp_precos_dryrun", destination="duckdb",
                            dataset_name="staging", dev_mode=True)
        pipe.run(anp_precos())
        with pipe.sql_client() as c:
            print(c.execute_sql(
                "SELECT indicador_id, count(*), avg(valor) AS preco_medio "
                "FROM indicadores_serie GROUP BY indicador_id ORDER BY count(*) DESC LIMIT 10"))
        return 0

    pipe = dlt.pipeline(pipeline_name="anp_precos", destination=s3tables_iceberg)
    info = pipe.run(anp_precos())
    print(info)
    return 0


if __name__ == "__main__":
    sys.exit(main())

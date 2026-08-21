"""DENATRAN/SENATRAN frota veicular por município → data_br.frota_municipio.

Carrega `s3://data-br-raw/denatran/raw/FrotapormunicipioetipoDezembro2024.xlsx`
(wide format: 1 row por (uf, município), 24 cols incluindo TOTAL +
20 tipos de veículo).

Faz pivot pra long:
  (ibge_code, uf, periodo='YYYY-MM', tipo_veiculo, combustivel=NULL, qtd)

Filename pattern: `Frotapormunicipioetipo{Mes}{YYYY}.xlsx`.

Mapeamento → FROTA_SCHEMA:
  ibge_code     ← resolve_ibge(uf, municipio) (slug match)
  uf            ← UF
  periodo       ← derivado do filename (YYYY-MM)
  tipo_veiculo  ← coluna xlsx (AUTOMOVEL, MOTOCICLETA, ...)
  combustivel   ← NULL (esse layout não tem; outro layout
                  `D_Frota_por_UF_Municipio_COMBUSTIVEL_*` cobre)
  qtd           ← valor da célula

Source: senatran.gov.br/estatistica → "frota por município".
Para histórico mensal, baixar `Frotapormunicipioetipo{Mes}{YYYY}.xlsx`
de cada release mensal.

Uso:
    AWS_PROFILE=<seu-perfil> python -m lakebrasil.pipelines.frota --no-fetch
    AWS_PROFILE=<seu-perfil> python -m lakebrasil.pipelines.frota --dry-run
"""
from __future__ import annotations

import argparse
import io
import re
import sys
import time
from collections.abc import Iterator

import dlt
import openpyxl

from lakebrasil.common.args import add_common_args
from lakebrasil.common.enrich import municipios_count, resolve_ibge
from lakebrasil.common.fetch import ensure_fetched
from lakebrasil.common.s3 import RAW_BUCKET, list_keys, s3_client
from lakebrasil.pipelines.destinations.s3tables import s3tables_iceberg

S3_PREFIX = "denatran/raw/"
# Match Frotapormunicipioetipo{Mes}{YYYY}.xlsx — sheet wide com 1 col
# por tipo de veículo. Outros formatos do DENATRAN (COMBUSTIVEL,
# ANO_FAB_MOD) têm layout diferente e ficam pra pipelines separados.
FILE_RE = re.compile(
    r"^Frotapormunicipioetipo([A-Za-z]+)(\d{4})\.xlsx$"
)

MES_MAP = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
    "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}

# Header da linha 4 do xlsx. UF + MUNICIPIO + TOTAL + 21 tipos.
SKIP_COLS = {"UF", "MUNICIPIO", "TOTAL"}


def _parse_filename(name: str) -> tuple[int, int] | None:
    m = FILE_RE.match(name)
    if not m:
        return None
    mes_str = m.group(1).lower().replace("ç", "c").replace("ã", "a")
    mes = MES_MAP.get(mes_str)
    if mes is None:
        return None
    return int(m.group(2)), mes


def _iter_xlsx(s3_key: str, ano: int, mes: int) -> Iterator[dict]:
    """Stream o xlsx (single sheet wide) → unpivot to long rows."""
    name = s3_key.rsplit("/", 1)[-1]
    print(f"  FROTA {name}: download S3", file=sys.stderr)
    s3 = s3_client()
    body = s3.get_object(Bucket=RAW_BUCKET, Key=s3_key)["Body"].read()
    wb = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    periodo = f"{ano:04d}-{mes:02d}"

    # Header é a linha 4 (rows 1-3 são título / SUM). Iter rows yield
    # tuples (UF, MUNICIPIO, TOTAL, AUTOMOVEL, BONDE, ...).
    header: list[str] = []
    emitted = 0
    skipped_ibge = 0
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if not row or all(v is None for v in row):
            continue
        if not header:
            # primeira row com 'UF' na coluna 0 é o header
            if row[0] == "UF" and len(row) > 5:
                header = [str(c).strip() if c else "" for c in row]
            continue
        uf = (str(row[0]) if row[0] else "").strip().upper()
        mun = (str(row[1]) if row[1] else "").strip()
        if not uf or not mun or uf == "UF":
            continue
        # Linhas de TOTAL/SUM no fim podem aparecer com mun vazio ou
        # UF == 'BRASIL' — pula.
        if uf in {"BRASIL", "TOTAL", ""}:
            continue
        ibge = resolve_ibge(uf, mun)
        if ibge is None:
            skipped_ibge += 1
            continue
        for col_idx, col_name in enumerate(header):
            if col_idx < 2 or col_name in SKIP_COLS or not col_name:
                continue
            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                continue
            try:
                qtd = int(val)
            except (ValueError, TypeError):
                continue
            if qtd == 0:
                continue
            yield {
                "ibge_code":    ibge,
                "uf":           uf,
                "periodo":      periodo,
                "tipo_veiculo": col_name,
                "combustivel":  None,
                "qtd":          qtd,
            }
            emitted += 1
    print(f"  FROTA {name}: emitted={emitted:,} skipped_ibge={skipped_ibge}",
          file=sys.stderr)


def _periods_loaded() -> set[str]:
    from pyiceberg.exceptions import NoSuchTableError

    from lakebrasil.loaders.iceberg import catalog
    try:
        t = catalog().load_table("data_br.frota_municipio")
    except NoSuchTableError:
        return set()
    if not t.metadata.snapshots:
        return set()
    arrow = t.scan(selected_fields=("periodo",)).to_arrow()
    return {p for p in arrow.column("periodo").to_pylist() if p}


def _build_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    add_common_args(p, table_default="frota_municipio")
    return p.parse_args()


def main() -> int:
    args = _build_args()

    if not args.no_fetch:
        try:
            ensure_fetched("denatran_*", refresh=args.refresh)
        except ValueError:
            print("  [warn] assumindo s3://data-br-raw/denatran/raw/ já populado.",
                  file=sys.stderr)

    municipios_count()  # warm dim
    already = set() if args.full_refresh else _periods_loaded()
    if already:
        print(f"  periodos já carregados: {sorted(already)}", file=sys.stderr)

    @dlt.resource(
        name="frota_municipio",
        primary_key=["ibge_code", "periodo", "tipo_veiculo"],
        write_disposition="append",
    )
    def frota() -> Iterator[dict]:
        for key in sorted(list_keys(S3_PREFIX)):
            name = key.rsplit("/", 1)[-1]
            parsed = _parse_filename(name)
            if parsed is None:
                continue
            ano, mes = parsed
            periodo = f"{ano:04d}-{mes:02d}"
            if periodo in already:
                continue
            t0 = time.monotonic()
            yield from _iter_xlsx(key, ano, mes)
            print(f"  FROTA {periodo}: pronto em {time.monotonic()-t0:.1f}s",
                  file=sys.stderr)

    if args.dry_run:
        pipe = dlt.pipeline(pipeline_name="frota_dryrun", destination="duckdb",
                            dataset_name="staging", dev_mode=True)
        pipe.run(frota())
        with pipe.sql_client() as c:
            print(c.execute_sql(
                "SELECT tipo_veiculo, count(*) AS munis, sum(qtd) AS total "
                "FROM frota_municipio GROUP BY tipo_veiculo ORDER BY total DESC LIMIT 15"))
        return 0

    pipe = dlt.pipeline(pipeline_name="frota", destination=s3tables_iceberg)
    info = pipe.run(frota())
    print(info)
    return 0


if __name__ == "__main__":
    sys.exit(main())

"""SINISA — Sistema Nacional Informações sobre Saneamento Básico → indicadores_serie.

Sucessor do SNIS (que foi descontinuado em 2023). Cobre 5 módulos:

  AGUA      Abastecimento (35 IAG + 22 IFA = 57 indicadores)
  ESGOTO    Esgotamento  (20 IES + 22 IFE = 42 indicadores)
  RESIDUOS  Resíduos sólidos (IFR + IRS)
  PLUVIAIS  Águas pluviais   (IGE + IFD + IAP + IGR)
  GESTAO    Gestão municipal (OGM*)

Estrutura raw em `s3://data-br-raw/sinisa/raw/`:
  sinisa_agua_2024.zip      → SINISA_Resultados_Ref2024/.../Indicadores_Base Municipal*.xlsx
  sinisa_esgoto_2024.zip    → Esgoto - Base Municipal/SINISA_ESGOTO_Indicadores*.xlsx
  sinisa_residuos_2024.zip  → SINISA_RESIDUOS_planilhas_2024/SINISA_RESIDUOS_Indicadores_2024.xlsx
  sinisa_pluviais_2024.zip  → SINISA_AGUASPLUVIAIS_Indicadores_2024.zip → *.xlsx
  sinisa_gestao_2024.xlsx   → direto

XLSX (não .xls como SNIS), IBGE 7-dígitos direto (sem crosswalk co6→ibge).
Cada módulo tem layout específico de header/code rows — ver MODULOS abaixo.

Output em data_br.indicadores_serie:
  fonte         = 'sinisa'
  indicador_id  = 'sinisa.{CODE}'    (ex: sinisa.IAG0001, sinisa.IES0001)
  periodo       = '2024'             (ano de referência)
  unidade       = row N+1 from xlsx (ex: 'Percentual', 'R$/m³')
  valor_texto   = nome humano do indicador

Aliases SNIS-compatíveis emitidos em paralelo (fonte='snis') pros 6
indicadores que SNIS já cobria — permite UNION com histórico SNIS
2020-2022 transparente pra dashboards.

Uso:
    AWS_PROFILE=<seu-perfil> python -m lakebrasil.pipelines.sinisa --no-fetch
    AWS_PROFILE=<seu-perfil> python -m lakebrasil.pipelines.sinisa --modulo AGUA --no-fetch
"""
from __future__ import annotations

import argparse
import io
import os
import re
import sys
import tempfile
import time
import zipfile
from collections.abc import Iterator
from dataclasses import dataclass

import dlt
import openpyxl

from lakebrasil.common.args import add_common_args
from lakebrasil.common.fetch import ensure_fetched
from lakebrasil.common.incremental import loaded_triples
from lakebrasil.common.s3 import get_object_bytes
from lakebrasil.pipelines.destinations.s3tables import s3tables_iceberg

S3_PREFIX = "sinisa/raw/"
PERIODO = "2024"  # ano de referência dos dados (publicado dez/2025)
CODE_RE = re.compile(r"^[A-Z]{2,4}\d{3,5}\s*$")


@dataclass(frozen=True)
class Modulo:
    """Layout config pra parser de cada módulo SINISA."""
    label: str             # AGUA/ESGOTO/RESIDUOS/PLUVIAIS/GESTAO
    s3_key: str            # arquivo no raw bucket
    sheet_substr: str      # substring que identifica a sheet de indicadores
    xlsx_substr: str       # substring que identifica o xlsx dentro do zip (ou "" se zip→raw é o xlsx)
    inner_zip_substr: str  # zip aninhado dentro do zip principal (PLUVIAIS), "" se não tem
    hname_row: int         # linha do nome humano do indicador (1-indexed)
    code_row: int          # linha do código (IAG0001, etc.)
    unit_row: int | None   # linha da unidade (None se ausente)
    data_start: int        # primeira linha de dados
    ibge_col: int          # coluna do código IBGE 7-dig (1-indexed)
    prefixes: tuple[str, ...]  # quais prefixes de código emitir como indicadores


MODULOS: list[Modulo] = [
    Modulo("AGUA", "sinisa_agua_2024.zip",
           sheet_substr="Indicadores de Gestão",
           xlsx_substr="Indicadores_Base Municipal", inner_zip_substr="",
           hname_row=9, code_row=11, unit_row=10, data_start=12, ibge_col=1,
           prefixes=("IAG", "IFA")),
    Modulo("ESGOTO", "sinisa_esgoto_2024.zip",
           sheet_substr="Indicadores - Esgoto",
           xlsx_substr="Indicadores_Base Municipal", inner_zip_substr="",
           hname_row=9, code_row=11, unit_row=10, data_start=12, ibge_col=1,
           prefixes=("IES", "IFE")),
    Modulo("RESIDUOS", "sinisa_residuos_2024.zip",
           sheet_substr="Planilha_Indicadores_municipais",
           xlsx_substr="RESIDUOS_Indicadores", inner_zip_substr="",
           hname_row=12, code_row=11, unit_row=13, data_start=14, ibge_col=2,
           prefixes=("IFR", "IRS")),
    Modulo("PLUVIAIS", "sinisa_pluviais_2024.zip",
           sheet_substr="Indicadores por município",
           xlsx_substr="AGUASPLUVIAIS_Indicadores",
           inner_zip_substr="AGUASPLUVIAIS_Indicadores",  # zip dentro do zip
           hname_row=8, code_row=11, unit_row=None, data_start=12, ibge_col=1,
           prefixes=("IGE", "IFD", "IAP", "IGR")),
    Modulo("GESTAO", "sinisa_gestao_2024.xlsx",
           sheet_substr="GM",
           xlsx_substr="",  # já é xlsx direto
           inner_zip_substr="",
           hname_row=10, code_row=11, unit_row=None, data_start=15, ibge_col=2,
           prefixes=("OGM",)),
]

# Aliases SINISA → SNIS pros 6 indicadores que SNIS publicou 2020-2022.
# Emitidos como rows extras com fonte='snis' pra permitir UNION transparente
# entre série SNIS 2020-2022 e SINISA 2024+. Equivalência semântica:
#   IAG0001 (% pop total c/ rede água)  ≡ snis.IN055 (atendimento total água)
#   IAG0002 (% pop urbana c/ rede água) ≡ snis.IN023 (atendimento urbano água)
#   IAG2013 (perdas totais distribuição)≡ snis.IN049 (perdas distribuição)
#   IES0001 (% pop total c/ rede esgoto)≡ snis.IN056 (atendimento total esgoto)
#   IES2002 (esgoto coletado/água cons.) ≈ snis.IN015 (índice coleta esgoto)
#   IES2003 (esgoto tratado/água cons.)  ≈ snis.IN016 (índice tratamento esgoto)
SNIS_ALIAS: dict[str, str] = {
    "IAG0001": "IN055",
    "IAG0002": "IN023",
    "IAG2013": "IN049",
    "IES0001": "IN056",
    "IES2002": "IN015",
    "IES2003": "IN016",
}


def _load_xlsx(mod: Modulo) -> openpyxl.Workbook:
    """Baixa do S3 + extrai aninhado se necessário + abre xlsx em modo read-only."""
    raw = get_object_bytes(S3_PREFIX + mod.s3_key)
    # Se o arquivo já é xlsx direto (GESTAO), carrega diret.
    if mod.s3_key.endswith(".xlsx"):
        return openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    # Caso .zip: extrai o xlsx alvo
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        if mod.inner_zip_substr:
            # PLUVIAIS: zip dentro do zip
            inner_zip_name = next(
                (n for n in zf.namelist()
                 if mod.inner_zip_substr in n and n.endswith(".zip")),
                None,
            )
            if not inner_zip_name:
                raise FileNotFoundError(
                    f"{mod.label}: inner zip *{mod.inner_zip_substr}*.zip não encontrado")
            inner_bytes = zf.read(inner_zip_name)
            with zipfile.ZipFile(io.BytesIO(inner_bytes)) as zi:
                xlsx_name = next(
                    (n for n in zi.namelist()
                     if mod.xlsx_substr in n and n.endswith(".xlsx")),
                    None,
                )
                if not xlsx_name:
                    raise FileNotFoundError(
                        f"{mod.label}: xlsx *{mod.xlsx_substr}*.xlsx não encontrado")
                xlsx_bytes = zi.read(xlsx_name)
        else:
            xlsx_name = next(
                (n for n in zf.namelist()
                 if mod.xlsx_substr in n and n.endswith(".xlsx")),
                None,
            )
            if not xlsx_name:
                raise FileNotFoundError(
                    f"{mod.label}: xlsx *{mod.xlsx_substr}*.xlsx não encontrado")
            xlsx_bytes = zf.read(xlsx_name)
    # openpyxl read-only não aceita BytesIO direto pra .xlsx grande sem seek
    # confiável — escreve em tempfile.
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        tf.write(xlsx_bytes)
        tmp_path = tf.name
    try:
        return openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def _find_sheet(wb: openpyxl.Workbook, substr: str):
    for name in wb.sheetnames:
        if substr in name:
            return wb[name]
    raise KeyError(f"sheet *{substr}* não encontrada (sheets={wb.sheetnames})")


def _iter_modulo(mod: Modulo, ibge_to_uf: dict[int, str]) -> Iterator[dict]:
    """Yield 1 dict por (município, indicador) de um módulo SINISA.

    Implementação stream-only via iter_rows pra evitar quadrático
    do sh.cell(r,c) em read_only mode (cada cell() faz seek O(N)
    desde o início da sheet).
    """
    t0 = time.monotonic()
    wb = _load_xlsx(mod)
    sh = _find_sheet(wb, mod.sheet_substr)

    # 1. Lê só os headers (rows 1..data_start-1) em 1 pass sequencial.
    cols: list[tuple[int, str, str, str]] = []  # (col_idx_0based, code, hname, unit)
    hname_arr: list = []
    code_arr: list = []
    unit_arr: list = []
    max_pre = max(mod.hname_row, mod.code_row,
                  mod.unit_row or 0, mod.data_start - 1)
    for r_idx, row in enumerate(
        sh.iter_rows(min_row=1, max_row=max_pre, values_only=True), start=1
    ):
        if r_idx == mod.hname_row:
            hname_arr = list(row)
        if r_idx == mod.code_row:
            code_arr = list(row)
        if mod.unit_row and r_idx == mod.unit_row:
            unit_arr = list(row)
    for c, code in enumerate(code_arr):
        if not code:
            continue
        code_s = str(code).strip()
        if not CODE_RE.match(code_s):
            continue
        prefix = "".join(ch for ch in code_s if ch.isalpha())
        if prefix not in mod.prefixes:
            continue
        hname = (hname_arr[c] if c < len(hname_arr) else "") or ""
        unit = (unit_arr[c] if c < len(unit_arr) else "") or ""
        cols.append((c, code_s, str(hname).strip()[:120], str(unit).strip()[:30]))

    print(f"  SINISA {mod.label}: {len(cols)} indicador cols detectadas, "
          f"sweep data rows={mod.data_start}+ ({sh.max_row} total)",
          file=sys.stderr)

    ibge_col_0 = mod.ibge_col - 1  # iter_rows é 0-indexed
    emitted = 0
    aliases = 0
    skipped_ibge = 0
    skipped_value = 0

    # 2. Itera data rows sequencialmente em 1 pass.
    for row in sh.iter_rows(min_row=mod.data_start, values_only=True):
        if not row or ibge_col_0 >= len(row):
            continue
        ibge_raw = row[ibge_col_0]
        if not ibge_raw:
            continue
        ibge_s = str(ibge_raw).strip()
        if not ibge_s.isdigit():
            continue
        ibge = int(ibge_s)
        uf = ibge_to_uf.get(ibge)
        if uf is None:
            skipped_ibge += 1
            continue
        for col_idx, code, hname, unit in cols:
            if col_idx >= len(row):
                continue
            raw = row[col_idx]
            if raw is None:
                continue
            raw_s = str(raw).strip()
            if not raw_s or raw_s.lower() in ("null", "-", "nd", "n/a", "n.d."):
                continue
            if raw_s.startswith("Não calculado"):
                continue
            try:
                valor = float(raw_s.replace(",", "."))
            except ValueError:
                skipped_value += 1
                continue
            yield {
                "ibge_code":     ibge,
                "uf":            uf,
                "fonte":         "sinisa",
                "indicador_id":  f"sinisa.{code}",
                "periodo":       PERIODO,
                "valor":         valor,
                "valor_texto":   hname or None,
                "unidade":       unit or None,
                "fonte_arquivo": mod.s3_key,
            }
            emitted += 1
            if code in SNIS_ALIAS:
                yield {
                    "ibge_code":     ibge,
                    "uf":            uf,
                    "fonte":         "snis",
                    "indicador_id":  f"snis.{SNIS_ALIAS[code]}",
                    "periodo":       PERIODO,
                    "valor":         valor,
                    "valor_texto":   None,
                    "unidade":       unit or "percentual",
                    "fonte_arquivo": f"sinisa-alias:{mod.s3_key}",
                }
                aliases += 1
    wb.close()
    print(f"  SINISA {mod.label}: emitidos {emitted:,} + {aliases:,} aliases SNIS "
          f"em {time.monotonic()-t0:.1f}s (skip_ibge={skipped_ibge}, "
          f"skip_value={skipped_value})", file=sys.stderr)


def _build_ibge_to_uf() -> dict[int, str]:
    """ibge_code 7-dig → UF sigla. None pra filtrar + enriquecer rows."""
    from lakebrasil.loaders.iceberg import catalog
    arrow = catalog().load_table("data_br.municipios").scan(
        selected_fields=("ibge_code", "uf")
    ).to_arrow()
    out: dict[int, str] = {}
    for ibge, uf in zip(arrow.column("ibge_code").to_pylist(),
                        arrow.column("uf").to_pylist()):
        if ibge is not None:
            out[int(ibge)] = uf or "??"
    return out


def _build_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    p.add_argument("--modulo", action="append",
                   choices=[m.label for m in MODULOS],
                   help="Só processar módulos especificados (default: todos).")
    add_common_args(p, table_default="indicadores_serie")
    return p.parse_args()


def main() -> int:
    args = _build_args()

    if not args.no_fetch:
        try:
            ensure_fetched("sinisa_*", refresh=args.refresh)
        except ValueError:
            print("  [warn] assumindo s3://data-br-raw/sinisa/raw/ já populado.",
                  file=sys.stderr)

    ibge_to_uf = _build_ibge_to_uf()
    print(f"SINISA→ibge dim: {len(ibge_to_uf):,} munis válidos", file=sys.stderr)

    selected = [m for m in MODULOS if not args.modulo or m.label in args.modulo]
    # Filtra por fonte=sinisa OU snis (aliases) — push-down evita full
    # scan da tabela multi-fonte.
    already = set()
    if not args.full_refresh:
        already = (
            loaded_triples("indicadores_serie", "fonte", "indicador_id", "periodo",
                           fonte="sinisa")
            | loaded_triples("indicadores_serie", "fonte", "indicador_id", "periodo",
                             fonte="snis")
        )

    # 1 pipe.run() por módulo — commits granulares, resume safe se algo cair
    # no meio. Cada módulo gera ~150k-300k rows, dá ~30s-2min de load step.
    for mod in selected:
        def _make_resource(m: Modulo):
            @dlt.resource(
                name="indicadores_serie",
                primary_key=["ibge_code", "indicador_id", "periodo"],
                write_disposition="append",
            )
            def _r() -> Iterator[dict]:
                for rec in _iter_modulo(m, ibge_to_uf):
                    if (rec["fonte"], rec["indicador_id"], rec["periodo"]) in already:
                        continue
                    yield rec
            return _r

        if args.dry_run:
            pipe = dlt.pipeline(pipeline_name=f"sinisa_{mod.label.lower()}_dry",
                                destination="duckdb", dataset_name="staging",
                                dev_mode=True)
            pipe.run(_make_resource(mod)())
            with pipe.sql_client() as c:
                print(c.execute_sql(
                    "SELECT fonte, indicador_id, count(*) "
                    "FROM indicadores_serie GROUP BY fonte, indicador_id "
                    "ORDER BY count(*) DESC LIMIT 8"))
            continue

        pipe = dlt.pipeline(pipeline_name=f"sinisa_{mod.label.lower()}",
                            destination=s3tables_iceberg)
        t0 = time.monotonic()
        info = pipe.run(_make_resource(mod)())
        print(f"  SINISA {mod.label}: pipe.run() em {time.monotonic()-t0:.1f}s",
              file=sys.stderr)
        print(info)
    return 0


if __name__ == "__main__":
    sys.exit(main())
